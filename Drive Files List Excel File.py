import openpyxl as xl
import os.path, time
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from tkinter import *
import win32api
import webbrowser
import os


def main():
    def callback(url):
        webbrowser.open_new(url)

    def headers(sheet):
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        center = Alignment(horizontal='center', vertical='center')
        titleColor = PatternFill(start_color='000000',
                                 end_color='000000',
                                 fill_type='solid')
        headingColor = PatternFill(start_color='808080',
                                   end_color='808080',
                                   fill_type='solid')

        titleFont = Font(color="00FF00", size=20)

        cell = sheet.cell(1, 1)
        cell.value = "For PC/Excel Automation - softwares.rubick.org"
        cell.hyperlink = "https://softwares.rubick.org/"

        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
        cell.fill = titleColor
        cell.font = titleFont
        cell.alignment = center
        cell = sheet.cell(2, 1)
        cell.value = "S.No"
        cell.alignment = center
        cell.font = Font(bold=True)
        cell.fill = headingColor
        cell.border = thin_border
        cell = sheet.cell(2, 2)
        cell.value = "File Names"
        cell.alignment = center
        cell.font = Font(bold=True)
        cell.fill = headingColor
        cell.border = thin_border
        cell = sheet.cell(2, 3)
        cell.value = "File Type"
        cell.alignment = center
        cell.font = Font(bold=True)
        cell.fill = headingColor
        cell.border = thin_border
        cell = sheet.cell(2, 4)
        cell.value = "File Size (Mbs)"
        cell.alignment = center
        cell.font = Font(bold=True)
        cell.fill = headingColor
        cell.border = thin_border
        cell = sheet.cell(2, 5)
        cell.value = "File Modified"
        cell.alignment = center
        cell.font = Font(bold=True)
        cell.fill = headingColor
        cell.border = thin_border
        cell = sheet.cell(2, 6)
        cell.value = "File Created"
        cell.alignment = center
        cell.font = Font(bold=True)
        cell.fill = headingColor
        cell.border = thin_border
        cell = sheet.cell(2, 7)
        cell.value = "Location"
        cell.alignment = center
        cell.font = Font(bold=True)
        cell.fill = headingColor

    def footers(sheet, ws):
        center = Alignment(horizontal='center', vertical='center')
        titleColor = PatternFill(start_color='000000',
                                 end_color='000000',
                                 fill_type='solid')

        titleFont = Font(color="00FF00", size=20)

        cell = sheet.cell(sheet.max_row + 1, 1)
        cell.value = "For PC/Excel Automation - softwares.rubick.org"
        cell.hyperlink = "https://softwares.rubick.org/"
        lastRow = cell.row
        sheet.merge_cells(start_row=lastRow, start_column=1, end_row=lastRow, end_column=7)
        cell.fill = titleColor
        cell.font = titleFont
        cell.alignment = center

        dim_holder = DimensionHolder(worksheet=ws)

        dim_holder[get_column_letter(1)] = ColumnDimension(ws, min=1, max=1, width=8)
        dim_holder[get_column_letter(2)] = ColumnDimension(ws, min=2, max=2, width=40)
        dim_holder[get_column_letter(3)] = ColumnDimension(ws, min=3, max=3, width=8)
        dim_holder[get_column_letter(4)] = ColumnDimension(ws, min=4, max=4, width=12)
        dim_holder[get_column_letter(5)] = ColumnDimension(ws, min=5, max=5, width=30)
        dim_holder[get_column_letter(6)] = ColumnDimension(ws, min=6, max=6, width=30)
        dim_holder[get_column_letter(7)] = ColumnDimension(ws, min=7, max=7, width=175)

        ws.column_dimensions = dim_holder

    def changeColor():
        checkBoxState = str(var.get())

        if checkBoxState == "1":
            checkBox.configure(fg='blue')
            root.geometry("320x315")
            startButton.place(x=105, y=238)
            checkBoxDescriptionWarning.place(x=28, y=162)

        if checkBoxState == "0":
            checkBox.configure(fg='black')
            root.geometry("320x260")
            startButton.place(x=105, y=177)
            checkBoxDescriptionWarning.place(x=28, y=368)


    def generateExcel():
        checkBoxState = str(var.get())
        drive = str(variable.get())
        drive_or_drives = []
        try:
            if drive != "All Drives":
                Drive = drive
                saveLocation = Drive + Drive[0] + " Drive (List Of All Files).xlsx"
                drive_or_drives.append(Drive)
            else:
                desktop = os.path.expanduser("~\desktop\\")
                saveLocation = desktop + "All Drives (List Of All Files).xlsx"
                drives = win32api.GetLogicalDriveStrings()
                drives = drives.split('\000')[:-1]
                drive_or_drives = list(drives)

            wb = xl.Workbook()
            wb.save(saveLocation)
            wb.create_sheet("List Of Files")
            ws = wb['List Of Files']
            del wb['Sheet']
            sheet = wb['List Of Files']
            wb.save(saveLocation)

            print("")
            print("Output File will Be Save here: " + saveLocation)
            print("")
            print("Donot Close the Terminal. It'll be closed automatically once done")


            def pointRemover(value):
                return str(value).replace('.', '')


            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            resultsColor = PatternFill(start_color='FFFFFF',
                                  end_color='FFFFFF',
                                  fill_type='solid')
            locationColor = PatternFill(start_color='00FFFF',
                                  end_color='00FFFF',
                                  fill_type='solid')

            sheet_counter = 1
            sheet_serial = 1
            Serial = 0
            for Drive in drive_or_drives:
                for subdir, dirs, files in os.walk(Drive):
                    for file in (files):
                        filepath = subdir + os.sep + file
                        if sheet_serial == 1_048_575:
                            footers(sheet, ws)
                            sheet_serial = 1
                            sheet_counter += 1
                            if sheet_counter == 1:
                                wb.create_sheet("List Of Files")
                                ws = wb['List Of Files']
                                sheet = wb['List Of Files']
                                wb.save(saveLocation)
                            else:
                                wb.create_sheet("List Of Files (" + str(sheet_counter) + ")")
                                ws = wb["List Of Files (" + str(sheet_counter) + ")"]
                                sheet = wb["List Of Files (" + str(sheet_counter) + ")"]
                                wb.save(saveLocation)
                        if sheet_serial == 1:
                            headers(sheet)
                            sheet_serial = 2
                        else:
                            if filepath.__contains__("$RECYCLE.BIN"):
                                pass
                            else:
                                print(f"{file} - {filepath}")
                                Serial += 1
                                sheet_serial += 1
                                cell = sheet.cell(sheet_serial, 1)
                                cell.value = Serial - 2
                                cell.fill = resultsColor
                                cell.border = thin_border
                                cell = sheet.cell(sheet_serial, 2)
                                cell.value = file
                                fileTypes = os.path.splitext(file)[1]
                                cell.fill = resultsColor
                                cell.border = thin_border
                                cell = sheet.cell(sheet_serial, 3)
                                fileType = pointRemover(fileTypes)
                                cell.value = fileType
                                cell.fill = resultsColor
                                cell.border = thin_border
                                cell = sheet.cell(sheet_serial, 4)
                                try:
                                    size = os.path.getsize(filepath)
                                    cell.fill = resultsColor
                                    cell.border = thin_border
                                    excelFile = str(file)
                                    if excelFile.__contains__("Drive (List Of All Files).xlsx"):
                                        cell.value = ""
                                    else:
                                        sizeMB = float(size / 1000000)
                                        mb_Points_Control = round(sizeMB, 2)
                                        cell.value = mb_Points_Control
                                except FileNotFoundError:
                                    cell.fill = resultsColor
                                    cell.border = thin_border
                                    pass
                                cell = sheet.cell(sheet_serial, 5)
                                try:
                                    dateModified = time.ctime(os.path.getmtime(filepath))
                                    cell.value = dateModified
                                    cell.fill = resultsColor
                                    cell.border = thin_border
                                except FileNotFoundError:
                                    cell.fill = resultsColor
                                    cell.border = thin_border
                                cell.fill = resultsColor
                                cell.border = thin_border
                                cell = sheet.cell(sheet_serial, 6)
                                try:
                                    dateCreated = time.ctime(os.path.getctime(filepath))
                                    cell.value = dateCreated
                                    cell.fill = resultsColor
                                    cell.border = thin_border
                                except FileNotFoundError:
                                    cell.fill = resultsColor
                                    cell.border = thin_border
                                cell = sheet.cell(sheet_serial, 7)
                                cell.value = filepath
                                cell.fill = locationColor
                                cell.border = thin_border
                                if checkBoxState == "1":
                                    cell.hyperlink = "\\\\" + filepath
                                    cell.style = "Hyperlink"
                                    cell.border = thin_border

            footers(sheet, ws)
            wb.save(saveLocation)
            os.startfile(saveLocation)
        except PermissionError:
            input("Either Excel File is already opened or there is no Admin Rights Granted. If Open then Please Close the file then Press Enter To Start over or Close the window to Close the Program...")
            print("")


    drives = win32api.GetLogicalDriveStrings()
    drives = drives.split('\000')[:-1]
    drives = list(drives)
    drives.append("All Drives")


    root = Tk()
    root.resizable(0,0)
    root.iconbitmap('icon.ico')
    root.title('Drive File Lister - V-2.1')
    # root.geometry("300x285")
    root.geometry("320x260")
    root.configure(bg="white")

    variable = StringVar(root)
    variable.set(drives[1])
    var = IntVar()

    # myFont = font.Font(family='Helvetica', size=20, weight='bold')


    label = Label(root, text="Drive Files Lister", font=("arial black", 20, 'bold'))
    label.configure(foreground="brown")
    label.configure(bg="white")
    label.place(x=37, y=0)

    selectDriveLabel = Label(root, text="Select Drive", font=("Comic Sans MS", 22, 'bold'))
    selectDriveLabel.configure(bg="white")
    selectDriveLabel.place(x=18, y=50)

    option = OptionMenu(root, variable, *drives)
    option.configure(cursor="hand1")
    option.place(x=220, y=60)

    optionDescriptionWarning = Label(root, text="For All Drives or C Drive. Run App with Admin Rights", font=("Calibri", 12, 'italic'), wraplength=320, justify='left')
    optionDescriptionWarning.configure(bg="white")
    optionDescriptionWarning.configure(bd=2)
    optionDescriptionWarning.configure(foreground="red")
    optionDescriptionWarning.place(x=10, y=90)

    checkBox = Checkbutton(root, text="Hyperlinks", variable=var, font=("Berlin Sans FB Demi", 15), cursor="hand1", command=changeColor)
    checkBox.configure(bg="white")
    checkBox.configure(fg="black")
    checkBox.place(x=100, y=125)

    checkBoxDescriptionWarning = Label(root, text="Enabling Hyperlink will take 3x more time while generating the file and might not work for low end PCs", font=("Calibri", 12, 'italic'), wraplength=280, justify='left')
    checkBoxDescriptionWarning.configure(bg="white")
    checkBoxDescriptionWarning.configure(bd=2)
    checkBoxDescriptionWarning.configure(foreground="red")
    # checkBoxDescriptionWarning.place(x=8, y=138)

    startButton = Button(root, text="S T A R T", font=("Arial", 15, 'bold'), justify='center', command=generateExcel, cursor="hand1")
    startButton.configure(foreground="yellow")
    startButton.configure(bg="black")
    # startButton.place(x=100, y=208)
    startButton.place(x=105, y=177)

    footer = Label(root, text="softwares.rubick.org", font=(14), cursor="hand2")
    footer.bind("<Button-1>", lambda e: callback("http://softwares.rubick.org"))
    footer.configure(foreground="white")
    footer.configure(bg="black")
    footer.pack(side=BOTTOM)
    root.mainloop()


if __name__ == "__main__":
    main()